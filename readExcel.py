from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#Create a workbook object
#wb = Workbook()

#load existing spreadsheet
wb = load_workbook(r'C:\Users\DELL\python_projects\data.xlsx')

#Create a worksheet
ws = wb.active

#set a variable
#name = ws["A2"].value
#colur = ws["B2"].value

#name = ws["A"]
#colur = ws["B"]

#print from spreadsheet
print("Initializing..........")
#print(f'{"Name"}:{"Colour"}')
#print(f'{ws["A2"].value}:{ws["B2"].value}')
#print(f'{name}: {colur}')

#for loop
#for cell in name:
    #print(f'{cell.value}\n : {col.value}\n')
    #print(cell.value)

#loop all the items
range = ws['A2':'B10']


#loop all items ['A2':'B10']
for cell in range:
    for x in cell:
        print(x.value)



        