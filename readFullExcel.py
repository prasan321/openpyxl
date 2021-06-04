from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#load existing spreadsheet
wb = load_workbook(r'C:\Users\DELL\python_projects\data.xlsx')

#Create a worksheet
ws = wb.active

#print from spreadsheet
print("Initializing..........")

#loop all the items 

#for value in ws.iter_rows(min_row=1, max_row=8, min_col=1, max_col=2):
    #print(value)
#    for x,y in value:
#       print(f'{x.value}')

#column only
#for column in ws.iter_cols(min_row=1, max_row=8, min_col=1, max_col=2):
    #print(column) 
#    for x in column:
#        print(x.value)    

#best way to read

for row in range(1, ws.max_row + 1):
    for column in range(1, ws.max_column + 1 ):
            print(ws.cell(row, column).value, end=" ")
    print()